import csv
import io
import json
import logging
from typing import Any
from uuid import UUID

from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from base import services as base_services
from career.models import Career, CareerImportBatch, CareerImportError
from education_level.models import EducationLevel

logger = logging.getLogger(__name__)

SAMPLE_CSV_HEADERS = (
    "career_code",
    "career_name",
    "min_education_level",
    "max_education_level",
    "description",
    "is_active",
)
REQUIRED_IMPORT_HEADERS = {"career_code", "career_name", "min_education_level"}
HEADER_ALIASES = {
    "code": "career_code",
    "career_master_code": "career_code",
    "name": "career_name",
    "min_education_level_code": "min_education_level",
    "max_education_level_code": "max_education_level",
    "active": "is_active",
    "status": "is_active",
}


def career_base_queryset():
    return Career.objects.select_related(
        "created_by",
        "updated_by",
        "min_education_level",
        "max_education_level",
    )


def case_insensitive_code_exists(*, code: str, exclude_pk: UUID | None = None) -> bool:
    q = Career.objects.filter(career_code__iexact=code)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    return q.exists()


def _resolve_education_level_value(value):
    if value in (None, ""):
        return None
    raw = str(value).strip()
    if not raw:
        return None
    obj = EducationLevel.objects.filter(level_code__iexact=raw, deleted=False).first()
    if obj:
        return obj
    try:
        return EducationLevel.objects.filter(pk=raw, deleted=False).first()
    except Exception:
        return None


def _assert_education_level_mapping(*, min_education_level, max_education_level):
    if not min_education_level:
        raise ValidationError({"min_education_level": "This field is required."})
    if max_education_level is None:
        return
    if max_education_level.sequence_order < min_education_level.sequence_order:
        raise ValidationError(
            {
                "max_education_level": "Max education level cannot be below min education level."
            }
        )


def _recommendation_blocker(career: Career) -> str | None:
    for rel in career._meta.related_objects:
        model_name = rel.related_model.__name__.lower()
        if "recommend" not in model_name:
            continue
        accessor = rel.get_accessor_name()
        try:
            related = getattr(career, accessor)
        except Exception:
            continue
        if hasattr(related, "exists"):
            if related.exists():
                return rel.related_model.__name__
        elif related is not None:
            return rel.related_model.__name__
    return None


def blocking_foreign_key_usage(career: Career) -> str | None:
    for rel in career._meta.related_objects:
        accessor = rel.get_accessor_name()
        try:
            related = getattr(career, accessor)
        except Exception:
            continue
        if hasattr(related, "exists"):
            if related.exists():
                return rel.related_model.__name__
        elif related is not None:
            return rel.related_model.__name__
    return None


def assert_can_archive(career: Career):
    recommendation_blocker = _recommendation_blocker(career)
    if recommendation_blocker:
        raise ValidationError(
            f"Cannot archive: used in recommendations ({recommendation_blocker})."
        )
    blocker = blocking_foreign_key_usage(career)
    if blocker:
        raise ValidationError(f"Cannot archive: referenced by {blocker}.")


@transaction.atomic
def create_career(*, user, validated_data: dict) -> Career:
    instance = Career(**validated_data)
    instance.save(user=user)
    return instance


@transaction.atomic
def update_career(*, career: Career, user, validated_data: dict) -> Career:
    for k, v in validated_data.items():
        setattr(career, k, v)
    career.save(user=user)
    return career


@transaction.atomic
def soft_archive_career(*, career: Career, user) -> Career:
    assert_can_archive(career)
    return base_services.soft_delete(career, user=user)


archive_career = soft_archive_career


@transaction.atomic
def restore_career(*, career: Career, user) -> Career:
    return base_services.restore(career, user=user)


@transaction.atomic
def bulk_archive(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    qs = Career.objects.filter(id__in=ids, deleted=False)
    count = 0
    for c in qs.select_related("min_education_level", "max_education_level"):
        assert_can_archive(c)
        c.soft_delete(user=user)
        count += 1
    return count


@transaction.atomic
def bulk_restore(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    updated = Career.objects.filter(id__in=ids, deleted=True).update(
        deleted=False,
        deleted_at=None,
        deleted_by=None,
        updated_at=timezone.now(),
        updated_by=user,
    )
    return updated


@transaction.atomic
def set_active_status(*, career: Career, user, is_active: bool) -> Career:
    career.is_active = is_active
    career.save(user=user)
    return career


@transaction.atomic
def bulk_set_active(*, ids: list, user, is_active: bool) -> int:
    if not ids:
        return 0
    return Career.objects.filter(id__in=ids).update(
        is_active=is_active,
        updated_at=timezone.now(),
        updated_by=user,
    )


def dropdown_careers():
    return (
        Career.objects.filter(is_active=True, deleted=False)
        .select_related("min_education_level", "max_education_level")
        .only(
            "id",
            "career_code",
            "career_name",
            "min_education_level__id",
            "min_education_level__level_code",
            "min_education_level__display_name",
            "max_education_level__id",
            "max_education_level__level_code",
            "max_education_level__display_name",
        )
        .order_by("career_name")
    )


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k, v in row.items():
        kk = (str(k) if k is not None else "").strip()
        if not kk:
            continue
        key = HEADER_ALIASES.get(kk.lower(), kk.lower())
        out[key] = v

    if "career_code" in out and out["career_code"] not in (None, ""):
        out["career_code"] = str(out["career_code"]).strip().lower()
    if "career_name" in out and out["career_name"] not in (None, ""):
        out["career_name"] = str(out["career_name"]).strip()
    if "description" in out and out["description"] not in (None, ""):
        out["description"] = str(out["description"]).strip()
    if "is_active" in out and out["is_active"] not in ("", None):
        out["is_active"] = str(out["is_active"]).lower() in ("1", "true", "yes", "y")

    for key in ("min_education_level", "max_education_level"):
        if key in out and out[key] in ("", None):
            out[key] = None

    return out


def parse_import_file(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    if not uploaded:
        return [], ["No file uploaded."]
    name = (getattr(uploaded, "name", "") or "").lower()
    raw = uploaded.read()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return [], [
                    "Excel support requires openpyxl. Install openpyxl or upload CSV."
                ]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], ["Empty spreadsheet."]
            headers = []
            for h in header_row:
                raw_h = str(h).strip() if h is not None else ""
                if raw_h:
                    headers.append(HEADER_ALIASES.get(raw_h.lower(), raw_h.lower()))
                else:
                    headers.append("")
            missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in headers if h]))
            if missing:
                return [], [f"Missing required headers: {', '.join(missing)}"]
            out_rows = []
            for tup in rows_iter:
                if all(x is None or str(x).strip() == "" for x in tup):
                    continue
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = tup[i] if i < len(tup) else None
                out_rows.append(row)
            wb.close()
            return out_rows, errors
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], ["CSV has no header row."]
        normalized_headers = [
            HEADER_ALIASES.get(
                (str(h).strip().lower() if h else ""),
                (str(h).strip().lower() if h else ""),
            )
            for h in reader.fieldnames
        ]
        missing = sorted(
            REQUIRED_IMPORT_HEADERS - set([h for h in normalized_headers if h])
        )
        if missing:
            return [], [f"Missing required headers: {', '.join(missing)}"]
        out_rows = []
        for r in reader:
            if not any((v or "").strip() for v in r.values()):
                continue
            out_rows.append(dict(r))
        return out_rows, errors
    except Exception as e:
        logger.exception("parse_import_file failed")
        return [], [str(e)]


@transaction.atomic
def bulk_import_rows(
    *, user, rows: list[dict], serializer_class, context: dict
) -> CareerImportBatch:
    batch = CareerImportBatch.objects.create(
        created_by=user,
        total_rows=len(rows),
    )
    imported = 0
    errors: list[CareerImportError] = []
    seen_codes: set[str] = set()
    for idx, raw_row in enumerate(rows, start=1):
        row = dict(raw_row)
        try:
            row = normalize_import_row(row)
        except ValueError as e:
            errors.append(
                CareerImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=dict(raw_row) if isinstance(raw_row, dict) else {},
                )
            )
            continue
        row_code = (row.get("career_code") or "").strip().lower()
        if row_code:
            if row_code in seen_codes:
                errors.append(
                    CareerImportError(
                        batch=batch,
                        row_number=idx,
                        message=f"Duplicate career_code in upload: {row_code}"[:500],
                        row_data=row if isinstance(row, dict) else {},
                    )
                )
                continue
            seen_codes.add(row_code)
        min_edu = _resolve_education_level_value(row.get("min_education_level"))
        if min_edu is None:
            errors.append(
                CareerImportError(
                    batch=batch,
                    row_number=idx,
                    message="Invalid min_education_level (use level_code or UUID of active education level).",
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        max_edu = _resolve_education_level_value(row.get("max_education_level"))
        if row.get("max_education_level") not in (None, "") and max_edu is None:
            errors.append(
                CareerImportError(
                    batch=batch,
                    row_number=idx,
                    message="Invalid max_education_level (use level_code or UUID of active education level).",
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        try:
            _assert_education_level_mapping(
                min_education_level=min_edu, max_education_level=max_edu
            )
        except ValidationError as e:
            detail = e.detail
            msg = (
                json.dumps(detail)[:500]
                if isinstance(detail, (dict, list))
                else str(detail)[:500]
            )
            errors.append(
                CareerImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        row["min_education_level"] = str(min_edu.pk)
        row["max_education_level"] = str(max_edu.pk) if max_edu else None

        existing = None
        if row_code:
            existing = Career.objects.filter(career_code__iexact=row_code).first()

        ser = serializer_class(
            instance=existing, data=row, partial=bool(existing), context=context
        )
        if not ser.is_valid():
            msg = json.dumps(ser.errors)[:500]
            errors.append(
                CareerImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        try:
            with transaction.atomic():
                obj = ser.save()
                if getattr(obj, "deleted", False):
                    obj.deleted = False
                    obj.deleted_at = None
                    obj.deleted_by = None
                    obj.updated_by = user
                    obj.updated_at = timezone.now()
                    obj.save(
                        update_fields=[
                            "deleted",
                            "deleted_at",
                            "deleted_by",
                            "updated_by",
                            "updated_at",
                        ]
                    )
                imported += 1
        except ValidationError as e:
            detail = e.detail
            if isinstance(detail, (dict, list)):
                msg = json.dumps(detail)[:500]
            else:
                msg = str(detail)[:500]
            errors.append(
                CareerImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
        except Exception as e:
            logger.exception("bulk row %s", idx)
            errors.append(
                CareerImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
    if errors:
        CareerImportError.objects.bulk_create(errors)
    batch.imported_count = imported
    batch.failed_count = len(errors)
    batch.completed_at = timezone.now()
    batch.save(update_fields=["imported_count", "failed_count", "completed_at"])
    return batch


def bulk_import_careers(
    *, user, rows: list[dict], serializer_class, context: dict
) -> dict[str, Any]:
    batch = bulk_import_rows(
        user=user,
        rows=rows,
        serializer_class=serializer_class,
        context=context,
    )
    err_qs = CareerImportError.objects.filter(batch=batch).order_by("row_number")
    error_details = [
        {"row": e.row_number, "message": e.message, "row_data": e.row_data}
        for e in err_qs.iterator(chunk_size=200)
    ]
    return {
        "success_count": batch.imported_count,
        "error_count": batch.failed_count,
        "error_details": error_details,
        "batch_id": str(batch.id),
    }


def import_batches_queryset():
    return CareerImportBatch.objects.select_related("created_by").order_by(
        "-created_at"
    )


def import_errors_queryset(*, batch_id: UUID | None = None):
    qs = CareerImportError.objects.select_related("batch").order_by(
        "-batch__created_at", "row_number"
    )
    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    return qs


def error_report_csv_bytes(*, batch_id: UUID | None = None) -> tuple[str, bytes]:
    qs = import_errors_queryset(batch_id=batch_id)
    if batch_id is None:
        first = qs.first()
        if not first:
            return "career_import_errors_empty.csv", b"row_number,message\n"
        qs = import_errors_queryset(batch_id=first.batch_id)
        filename = f"career_import_errors_{first.batch_id}.csv"
    else:
        filename = f"career_import_errors_{batch_id}.csv"
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["row_number", "message", "row_data"])
    for er in qs.iterator(chunk_size=200):
        w.writerow([er.row_number, er.message, er.row_data])
    return filename, buf.getvalue().encode("utf-8")


def sample_csv_bytes() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SAMPLE_CSV_HEADERS)
    w.writerow(
        [
            "software_engineer",
            "Software Engineer",
            "higher_secondary",
            "graduation",
            "Build software systems",
            "1",
        ]
    )
    w.writerow(
        [
            "data_analyst",
            "Data Analyst",
            "graduation",
            "",
            "Analyze and interpret data",
            "1",
        ]
    )
    return buf.getvalue().encode("utf-8")
