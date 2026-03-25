import csv
import io
import json
import logging
from typing import Any
from uuid import UUID

from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from base import services as base_services
from skill.models import Skill, SkillImportBatch, SkillImportError, SkillType

logger = logging.getLogger(__name__)

SAMPLE_CSV_HEADERS = (
    "skill_code",
    "skill_name",
    "skill_type",
    "description",
    "is_active",
)
REQUIRED_IMPORT_HEADERS = {"skill_code", "skill_name", "skill_type"}
HEADER_ALIASES = {
    "code": "skill_code",
    "skill_master_code": "skill_code",
    "name": "skill_name",
    "type": "skill_type",
    "active": "is_active",
    "status": "is_active",
}


def skill_base_queryset():
    return Skill.objects.select_related("created_by", "updated_by")


def list_skills(*, include_archived: bool = False):
    qs = skill_base_queryset()
    if not include_archived:
        qs = qs.filter(is_archived=False)
    return qs.order_by("-created_at")


def get_skill(*, pk: UUID, include_archived: bool = False):
    qs = skill_base_queryset().filter(pk=pk)
    if not include_archived:
        qs = qs.filter(is_archived=False)
    return qs.first()


def filter_skills(
    queryset,
    *,
    is_active=None,
    skill_type=None,
):
    if is_active is not None:
        queryset = queryset.filter(is_active=is_active)
    if skill_type not in (None, ""):
        queryset = queryset.filter(skill_type=str(skill_type).strip().lower())
    return queryset


def case_insensitive_code_exists(*, code: str, exclude_pk: UUID | None = None) -> bool:
    q = Skill.objects.filter(skill_code__iexact=code)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    return q.exists()


def blocking_foreign_key_usage(skill: Skill) -> str | None:
    for rel in skill._meta.related_objects:
        accessor = rel.get_accessor_name()
        try:
            related = getattr(skill, accessor)
        except Exception:
            continue
        if hasattr(related, "exists"):
            if related.exists():
                return rel.related_model.__name__
        elif related is not None:
            return rel.related_model.__name__
    return None


def assert_can_archive(skill: Skill):
    blocker = blocking_foreign_key_usage(skill)
    if blocker:
        raise ValidationError(f"Cannot archive: referenced by {blocker}.")


def validate_skill_type(value: Any) -> str:
    raw = (value or "").strip().lower()
    allowed = {c for c, _ in SkillType.choices}
    if raw not in allowed:
        raise ValidationError({"skill_type": f"Invalid skill_type. Allowed: {', '.join(sorted(allowed))}."})
    return raw


def validate_skill_data(*, data: dict[str, Any], instance: Skill | None = None) -> dict[str, Any]:
    code = (data.get("skill_code") or "").strip()
    if not code:
        raise ValidationError({"skill_code": "This field may not be blank."})
    name = (data.get("skill_name") or "").strip()
    if not name:
        raise ValidationError({"skill_name": "This field may not be blank."})
    exclude_pk = instance.pk if instance and instance.pk else None
    if case_insensitive_code_exists(code=code, exclude_pk=exclude_pk):
        raise ValidationError({"skill_code": "Skill code must be unique (case-insensitive)."})
    st = validate_skill_type(data.get("skill_type"))
    return {
        "skill_code": code,
        "skill_name": name,
        "skill_type": st,
        "description": (data.get("description") or "").strip(),
        "is_active": bool(data.get("is_active", True)),
    }


@transaction.atomic
def create_skill(*, user, validated_data: dict) -> Skill:
    instance = Skill(**validated_data)
    instance.save(user=user)
    return instance


@transaction.atomic
def update_skill(*, skill: Skill, user, validated_data: dict) -> Skill:
    for k, v in validated_data.items():
        setattr(skill, k, v)
    skill.save(user=user)
    return skill


@transaction.atomic
def soft_archive_skill(*, skill: Skill, user) -> Skill:
    assert_can_archive(skill)
    return base_services.soft_delete(skill, user=user)


archive_skill = soft_archive_skill


@transaction.atomic
def restore_skill(*, skill: Skill, user) -> Skill:
    return base_services.restore(skill, user=user)


@transaction.atomic
def bulk_archive(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    qs = Skill.objects.filter(id__in=ids, is_archived=False)
    count = 0
    for s in qs:
        assert_can_archive(s)
        s.is_archived = True
        s.save(user=user)
        count += 1
    return count


@transaction.atomic
def bulk_restore(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    updated = Skill.objects.filter(id__in=ids, is_archived=True).update(
        is_archived=False,
        updated_at=timezone.now(),
        updated_by=user,
    )
    return updated


@transaction.atomic
def set_active_status(*, skill: Skill, user, is_active: bool) -> Skill:
    skill.is_active = is_active
    skill.save(user=user)
    return skill


@transaction.atomic
def bulk_set_active(*, ids: list, user, is_active: bool) -> int:
    if not ids:
        return 0
    return Skill.objects.filter(id__in=ids).update(
        is_active=is_active,
        updated_at=timezone.now(),
        updated_by=user,
    )


def dropdown_skills():
    return (
        Skill.objects.filter(is_active=True, is_archived=False)
        .only("id", "skill_code", "skill_name", "skill_type")
        .order_by("skill_name")
    )


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k, v in row.items():
        kk = (str(k) if k is not None else "").strip()
        if not kk:
            continue
        key = HEADER_ALIASES.get(kk.lower(), kk.lower())
        out[key] = v

    if "skill_code" in out and out["skill_code"] not in (None, ""):
        out["skill_code"] = str(out["skill_code"]).strip()

    if "skill_type" in out and out["skill_type"] not in (None, ""):
        out["skill_type"] = str(out["skill_type"]).strip().lower()

    if "skill_name" in out and out["skill_name"] not in (None, ""):
        out["skill_name"] = str(out["skill_name"]).strip()

    if "description" in out and out["description"] not in (None, ""):
        out["description"] = str(out["description"]).strip()

    if "is_active" in out and out["is_active"] not in ("", None):
        out["is_active"] = str(out["is_active"]).lower() in ("1", "true", "yes", "y")

    return out


def parse_import_file(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    if not uploaded:
        return [], ["No file uploaded."]
    name = (getattr(uploaded, "name", "") or "").lower()
    raw = uploaded.read()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return [], ["Excel support requires openpyxl. Install openpyxl or upload CSV."]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], ["Empty spreadsheet."]
            headers = []
            for h in header_row:
                raw_h = str(h).strip() if h is not None else ""
                if raw_h:
                    headers.append(HEADER_ALIASES.get(raw_h.lower(), raw_h.lower()))
                else:
                    headers.append("")
            missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in headers if h]))
            if missing:
                return [], [f"Missing required headers: {', '.join(missing)}"]
            out_rows = []
            for tup in rows_iter:
                if all(x is None or str(x).strip() == "" for x in tup):
                    continue
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = tup[i] if i < len(tup) else None
                out_rows.append(row)
            wb.close()
            return out_rows, errors
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], ["CSV has no header row."]
        normalized_headers = [
            HEADER_ALIASES.get((str(h).strip().lower() if h else ""), (str(h).strip().lower() if h else ""))
            for h in reader.fieldnames
        ]
        missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in normalized_headers if h]))
        if missing:
            return [], [f"Missing required headers: {', '.join(missing)}"]
        out_rows = []
        for r in reader:
            if not any((v or "").strip() for v in r.values()):
                continue
            out_rows.append(dict(r))
        return out_rows, errors
    except Exception as e:
        logger.exception("parse_import_file failed")
        return [], [str(e)]


@transaction.atomic
def bulk_import_rows(*, user, rows: list[dict], serializer_class, context: dict) -> SkillImportBatch:
    batch = SkillImportBatch.objects.create(
        created_by=user,
        total_rows=len(rows),
    )
    imported = 0
    errors: list[SkillImportError] = []
    seen_codes: set[str] = set()
    for idx, raw_row in enumerate(rows, start=1):
        row = dict(raw_row)
        try:
            row = normalize_import_row(row)
        except ValueError as e:
            errors.append(
                SkillImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=dict(raw_row) if isinstance(raw_row, dict) else {},
                )
            )
            continue
        row_code = (row.get("skill_code") or "").strip().lower()
        if row_code:
            if row_code in seen_codes:
                errors.append(
                    SkillImportError(
                        batch=batch,
                        row_number=idx,
                        message=f"Duplicate skill_code in upload: {row_code}"[:500],
                        row_data=row if isinstance(row, dict) else {},
                    )
                )
                continue
            seen_codes.add(row_code)
        existing = None
        if row_code:
            existing = Skill.objects.filter(skill_code__iexact=row_code).first()
        ser = serializer_class(instance=existing, data=row, partial=bool(existing), context=context)
        if not ser.is_valid():
            msg = json.dumps(ser.errors)[:500]
            errors.append(
                SkillImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        try:
            with transaction.atomic():
                obj = ser.save()
                if getattr(obj, "is_archived", False):
                    obj.is_archived = False
                    obj.updated_by = user
                    obj.updated_at = timezone.now()
                    obj.save(update_fields=["is_archived", "updated_by", "updated_at"])
                imported += 1
        except ValidationError as e:
            detail = e.detail
            if isinstance(detail, (dict, list)):
                msg = json.dumps(detail)[:500]
            else:
                msg = str(detail)[:500]
            errors.append(
                SkillImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
        except Exception as e:
            logger.exception("bulk row %s", idx)
            errors.append(
                SkillImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
    if errors:
        SkillImportError.objects.bulk_create(errors)
    batch.imported_count = imported
    batch.failed_count = len(errors)
    batch.completed_at = timezone.now()
    batch.save(update_fields=["imported_count", "failed_count", "completed_at"])
    return batch


def bulk_import_skills(*, user, rows: list[dict], serializer_class, context: dict) -> dict[str, Any]:
    batch = bulk_import_rows(
        user=user,
        rows=rows,
        serializer_class=serializer_class,
        context=context,
    )
    err_qs = SkillImportError.objects.filter(batch=batch).order_by("row_number")
    error_details = [{"row": e.row_number, "message": e.message, "row_data": e.row_data} for e in err_qs.iterator(chunk_size=200)]
    return {
        "success_count": batch.imported_count,
        "error_count": batch.failed_count,
        "error_details": error_details,
        "batch_id": str(batch.id),
    }


def import_batches_queryset():
    return SkillImportBatch.objects.select_related("created_by").order_by("-created_at")


def import_errors_queryset(*, batch_id: UUID | None = None):
    qs = SkillImportError.objects.select_related("batch").order_by("-batch__created_at", "row_number")
    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    return qs


def error_report_csv_bytes(*, batch_id: UUID | None = None) -> tuple[str, bytes]:
    qs = import_errors_queryset(batch_id=batch_id)
    if batch_id is None:
        first = qs.first()
        if not first:
            return "skill_import_errors_empty.csv", b"row_number,message\n"
        qs = import_errors_queryset(batch_id=first.batch_id)
        filename = f"skill_import_errors_{first.batch_id}.csv"
    else:
        filename = f"skill_import_errors_{batch_id}.csv"
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["row_number", "message", "row_data"])
    for er in qs.iterator(chunk_size=200):
        w.writerow([er.row_number, er.message, er.row_data])
    return filename, buf.getvalue().encode("utf-8")


def sample_csv_bytes() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SAMPLE_CSV_HEADERS)
    w.writerow(["python", "Python", "technical", "Programming language", "1"])
    w.writerow(["communication", "Communication", "soft", "Verbal & written communication", "1"])
    return buf.getvalue().encode("utf-8")

