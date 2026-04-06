import csv
import io
import json
import logging
from typing import Any
from uuid import UUID

from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from stream.models import Stream, StreamImportBatch, StreamImportError

logger = logging.getLogger(__name__)

SAMPLE_CSV_HEADERS = (
    "stream_code",
    "stream_name",
    "sequence_order",
    "parent_safe_label",
    "traditional_equivalent",
    "description",
    "education_level",
    "is_active",
)
REQUIRED_IMPORT_HEADERS = {
    "stream_code",
    "stream_name",
    "sequence_order",
    "parent_safe_label",
}
HEADER_ALIASES = {
    "code": "stream_code",
    "stream_master_code": "stream_code",
    "name": "stream_name",
    "order": "sequence_order",
    "sequence": "sequence_order",
    "parent_safe": "parent_safe_label",
    "education_level_code": "education_level",
    "level_code": "education_level",
    "active": "is_active",
    "status": "is_active",
}


def stream_base_queryset():
    return Stream.objects.select_related("created_by", "updated_by", "education_level")


def case_insensitive_code_exists(*, code: str, exclude_pk: UUID | None = None) -> bool:
    q = Stream.objects.filter(stream_code__iexact=code)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    return q.exists()


def sequence_exists(*, sequence_order: int, exclude_pk: UUID | None = None) -> bool:
    q = Stream.objects.filter(sequence_order=sequence_order)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    return q.exists()


def blocking_foreign_key_usage(stream: Stream) -> str | None:
    for rel in stream._meta.related_objects:
        accessor = rel.get_accessor_name()
        try:
            related = getattr(stream, accessor)
        except Exception:
            continue
        if hasattr(related, "exists"):
            if related.exists():
                return rel.related_model.__name__
        elif related is not None:
            return rel.related_model.__name__
    return None


@transaction.atomic
def create_stream(*, user, validated_data: dict) -> Stream:
    instance = Stream(**validated_data)
    instance.save(user=user)
    return instance


@transaction.atomic
def update_stream(*, stream: Stream, user, validated_data: dict) -> Stream:
    for k, v in validated_data.items():
        setattr(stream, k, v)
    stream.save(user=user)
    return stream


def assert_can_archive(stream: Stream):
    blocker = blocking_foreign_key_usage(stream)
    if blocker:
        raise ValidationError(f"Cannot archive: referenced by {blocker}.")


@transaction.atomic
def archive_stream(*, stream: Stream, user) -> Stream:
    assert_can_archive(stream)
    stream.soft_delete(user=user)
    return stream


@transaction.atomic
def restore_stream(*, stream: Stream, user) -> Stream:
    stream.deleted = False
    stream.deleted_at = None
    stream.deleted_by = None
    stream.updated_at = timezone.now()
    stream.updated_by = user
    stream.save(
        update_fields=[
            "deleted",
            "deleted_at",
            "deleted_by",
            "updated_at",
            "updated_by",
        ]
    )
    return stream


@transaction.atomic
def bulk_archive(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    qs = Stream.objects.filter(id__in=ids, deleted=False)
    count = 0
    for stream in qs:
        assert_can_archive(stream)
        stream.soft_delete(user=user)
        count += 1
    return count


@transaction.atomic
def bulk_restore(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    return Stream.objects.filter(id__in=ids, deleted=True).update(
        deleted=False,
        deleted_at=None,
        deleted_by=None,
        updated_at=timezone.now(),
        updated_by=user,
    )


@transaction.atomic
def set_active_status(*, stream: Stream, user, is_active: bool) -> Stream:
    stream.is_active = is_active
    stream.save(user=user)
    return stream


@transaction.atomic
def bulk_set_active(*, ids: list, user, is_active: bool) -> int:
    if not ids:
        return 0
    return Stream.objects.filter(id__in=ids).update(
        is_active=is_active,
        updated_at=timezone.now(),
        updated_by=user,
    )


def dropdown_streams():
    return (
        Stream.objects.filter(is_active=True, deleted=False)
        .select_related("education_level")
        .only(
            "id",
            "stream_code",
            "stream_name",
            "sequence_order",
            "education_level__id",
            "education_level__level_code",
            "education_level__display_name",
        )
        .order_by("sequence_order", "stream_name")
    )


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {}
    for k, v in row.items():
        kk = (str(k) if k is not None else "").strip()
        if not kk:
            continue
        key = HEADER_ALIASES.get(kk.lower(), kk.lower())
        out[key] = v

    if "sequence_order" in out and out["sequence_order"] not in ("", None):
        try:
            out["sequence_order"] = int(float(str(out["sequence_order"]).strip()))
        except (TypeError, ValueError):
            raise ValueError("Invalid sequence_order")

    for b in ("parent_safe_label", "is_active"):
        if b in out and out[b] not in ("", None):
            out[b] = str(out[b]).lower() in ("1", "true", "yes", "y")

    if "education_level" in out and out["education_level"] in ("", None):
        out["education_level"] = None

    return out


def parse_import_file(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    if not uploaded:
        return [], ["No file uploaded."]
    name = (getattr(uploaded, "name", "") or "").lower()
    raw = uploaded.read()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return [], [
                    "Excel support requires openpyxl. Install openpyxl or upload CSV."
                ]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], ["Empty spreadsheet."]
            headers = []
            for h in header_row:
                raw_h = str(h).strip() if h is not None else ""
                if raw_h:
                    headers.append(HEADER_ALIASES.get(raw_h.lower(), raw_h.lower()))
                else:
                    headers.append("")
            missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in headers if h]))
            if missing:
                return [], [f"Missing required headers: {', '.join(missing)}"]
            out_rows = []
            for tup in rows_iter:
                if all(x is None or str(x).strip() == "" for x in tup):
                    continue
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = tup[i] if i < len(tup) else None
                out_rows.append(row)
            wb.close()
            return out_rows, errors
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], ["CSV has no header row."]
        normalized_headers = [
            HEADER_ALIASES.get(
                (str(h).strip().lower() if h else ""),
                (str(h).strip().lower() if h else ""),
            )
            for h in reader.fieldnames
        ]
        missing = sorted(
            REQUIRED_IMPORT_HEADERS - set([h for h in normalized_headers if h])
        )
        if missing:
            return [], [f"Missing required headers: {', '.join(missing)}"]
        out_rows = []
        for r in reader:
            if not any((v or "").strip() for v in r.values()):
                continue
            out_rows.append(dict(r))
        return out_rows, errors
    except Exception as e:
        logger.exception("parse_import_file failed")
        return [], [str(e)]


def _resolve_education_level_value(value):
    if value in (None, ""):
        return None
    from education_level.models import EducationLevel

    raw = str(value).strip()
    if not raw:
        return None
    obj = EducationLevel.objects.filter(level_code__iexact=raw, deleted=False).first()
    if obj:
        return obj
    try:
        return EducationLevel.objects.filter(pk=raw, deleted=False).first()
    except Exception:
        return None


@transaction.atomic
def bulk_import_rows(
    *, user, rows: list[dict], serializer_class, context: dict
) -> StreamImportBatch:
    batch = StreamImportBatch.objects.create(
        created_by=user,
        total_rows=len(rows),
    )
    imported = 0
    errors: list[StreamImportError] = []
    for idx, raw_row in enumerate(rows, start=1):
        row = dict(raw_row)
        try:
            row = normalize_import_row(row)
        except ValueError as e:
            errors.append(
                StreamImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=dict(raw_row) if isinstance(raw_row, dict) else {},
                )
            )
            continue
        if "education_level" in row:
            edu = _resolve_education_level_value(row.get("education_level"))
            if row.get("education_level") not in (None, "") and edu is None:
                errors.append(
                    StreamImportError(
                        batch=batch,
                        row_number=idx,
                        message="Invalid education_level (use level_code or UUID of active education level).",
                        row_data=row if isinstance(row, dict) else {},
                    )
                )
                continue
            row["education_level"] = str(edu.pk) if edu else None
        existing = None
        stream_code = (row.get("stream_code") or "").strip().lower()
        if stream_code:
            existing = Stream.objects.filter(stream_code__iexact=stream_code).first()
        ser = serializer_class(
            instance=existing, data=row, partial=bool(existing), context=context
        )
        if not ser.is_valid():
            msg = json.dumps(ser.errors)[:500]
            errors.append(
                StreamImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        try:
            with transaction.atomic():
                obj = ser.save()
                if getattr(obj, "deleted", False):
                    obj.deleted = False
                    obj.deleted_at = None
                    obj.deleted_by = None
                    obj.updated_by = user
                    obj.updated_at = timezone.now()
                    obj.save(
                        update_fields=[
                            "deleted",
                            "deleted_at",
                            "deleted_by",
                            "updated_by",
                            "updated_at",
                        ]
                    )
                imported += 1
        except ValidationError as e:
            detail = e.detail
            if isinstance(detail, (dict, list)):
                msg = json.dumps(detail)[:500]
            else:
                msg = str(detail)[:500]
            errors.append(
                StreamImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
        except Exception as e:
            logger.exception("bulk row %s", idx)
            errors.append(
                StreamImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
    if errors:
        StreamImportError.objects.bulk_create(errors)
    batch.imported_count = imported
    batch.failed_count = len(errors)
    batch.completed_at = timezone.now()
    batch.save(update_fields=["imported_count", "failed_count", "completed_at"])
    return batch


def bulk_import_streams(
    *, user, rows: list[dict], serializer_class, context: dict
) -> dict[str, Any]:
    batch = bulk_import_rows(
        user=user,
        rows=rows,
        serializer_class=serializer_class,
        context=context,
    )
    err_qs = StreamImportError.objects.filter(batch=batch).order_by("row_number")
    error_details = [
        {"row": e.row_number, "message": e.message, "row_data": e.row_data}
        for e in err_qs.iterator(chunk_size=200)
    ]
    return {
        "success_count": batch.imported_count,
        "error_count": batch.failed_count,
        "error_details": error_details,
        "batch_id": str(batch.id),
    }


def import_batches_queryset():
    return StreamImportBatch.objects.select_related("created_by").order_by(
        "-created_at"
    )


def import_errors_queryset(*, batch_id: UUID | None = None):
    qs = StreamImportError.objects.select_related("batch").order_by(
        "-batch__created_at", "row_number"
    )
    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    return qs


def error_report_csv_bytes(*, batch_id: UUID | None = None) -> tuple[str, bytes]:
    qs = import_errors_queryset(batch_id=batch_id)
    if batch_id is None:
        first = qs.first()
        if not first:
            return "stream_import_errors_empty.csv", b"row_number,message\n"
        qs = import_errors_queryset(batch_id=first.batch_id)
        filename = f"stream_import_errors_{first.batch_id}.csv"
    else:
        filename = f"stream_import_errors_{batch_id}.csv"
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["row_number", "message", "row_data"])
    for er in qs.iterator(chunk_size=200):
        w.writerow([er.row_number, er.message, er.row_data])
    return filename, buf.getvalue().encode("utf-8")


def sample_csv_bytes() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SAMPLE_CSV_HEADERS)
    w.writerow(
        [
            "science",
            "Science",
            "1",
            "1",
            "General Science",
            "Physics/Chemistry/Biology path",
            "secondary",
            "1",
        ]
    )
    w.writerow(
        [
            "commerce",
            "Commerce",
            "2",
            "1",
            "Business Studies",
            "Accounts/Economics path",
            "higher_secondary",
            "1",
        ]
    )
    return buf.getvalue().encode("utf-8")
