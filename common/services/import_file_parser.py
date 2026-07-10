"""Shared CSV/Excel import parsing for master data services."""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


def parse_validated_import_file(
    uploaded,
    *,
    header_aliases: dict[str, str],
    required_headers: set[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parse upload with header alias normalization and required-header checks.
    Used by education_level and stream import flows.
    """
    errors: list[str] = []
    if not uploaded:
        return [], ["No file uploaded."]
    name = (getattr(uploaded, "name", "") or "").lower()
    raw = uploaded.read()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return [], [
                    "Excel support requires openpyxl. Install openpyxl or upload CSV."
                ]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], ["Empty spreadsheet."]
            headers = []
            for h in header_row:
                raw_h = str(h).strip() if h is not None else ""
                if raw_h:
                    headers.append(header_aliases.get(raw_h.lower(), raw_h.lower()))
                else:
                    headers.append("")
            missing = sorted(required_headers - set(h for h in headers if h))
            if missing:
                return [], [f"Missing required headers: {', '.join(missing)}"]
            out_rows = []
            for tup in rows_iter:
                if all(x is None or str(x).strip() == "" for x in tup):
                    continue
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = tup[i] if i < len(tup) else None
                out_rows.append(row)
            wb.close()
            return out_rows, errors
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], ["CSV has no header row."]
        normalized_headers = [
            header_aliases.get(
                (str(h).strip().lower() if h else ""),
                (str(h).strip().lower() if h else ""),
            )
            for h in reader.fieldnames
        ]
        missing = sorted(required_headers - set(h for h in normalized_headers if h))
        if missing:
            return [], [f"Missing required headers: {', '.join(missing)}"]
        out_rows = []
        for r in reader:
            if not any((v or "").strip() for v in r.values()):
                continue
            out_rows.append(dict(r))
        return out_rows, errors
    except Exception as e:
        logger.exception("parse_import_file failed")
        return [], [str(e)]


def build_import_parser(header_aliases: dict[str, str], required_headers: set[str]):
    def parse_import_file(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
        return parse_validated_import_file(
            uploaded,
            header_aliases=header_aliases,
            required_headers=required_headers,
        )

    return parse_import_file
