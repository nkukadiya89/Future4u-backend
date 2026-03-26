import csv
import io
import json
import logging
from typing import Any
from uuid import UUID

from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from base.services import BaseService
from stream_domain_mapping.models import (
    StreamDomainMapping,
    StreamDomainMappingImportBatch,
    StreamDomainMappingImportError,
)

logger = logging.getLogger(__name__)

SAMPLE_CSV_HEADERS = ("stream_code", "domain_code", "weight_score", "is_primary", "is_active")
REQUIRED_IMPORT_HEADERS = {"stream_code", "domain_code", "weight_score"}
HEADER_ALIASES = {
    "stream": "stream_code",
    "domain": "domain_code",
    "score": "weight_score",
    "primary": "is_primary",
    "active": "is_active",
}


def mapping_base_queryset():
    return StreamDomainMapping.objects.select_related("stream", "domain", "created_by", "updated_by")


def _to_bool(value, *, default=False):
    if value in (None, ""):
        return default
    return str(value).strip().lower() in ("1", "true", "yes", "y")


def _normalize_code(value):
    return (str(value).strip().lower() if value is not None else "")


def _resolve_stream(stream_code):
    if stream_code in (None, ""):
        return None
    from stream.models import Stream

    return Stream.objects.filter(stream_code__iexact=str(stream_code).strip(), deleted=False).first()


def _resolve_domain(domain_code):
    if domain_code in (None, ""):
        return None
    from domain.models import Domain

    return Domain.objects.filter(domain_code__iexact=str(domain_code).strip(), deleted=False).first()


def pair_exists(*, stream_id, domain_id, exclude_pk: UUID | None = None) -> bool:
    q = StreamDomainMapping.objects.filter(stream_id=stream_id, domain_id=domain_id)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    return q.exists()


def blocking_foreign_key_usage(mapping: StreamDomainMapping) -> str | None:
    for rel in mapping._meta.related_objects:
        accessor = rel.get_accessor_name()
        try:
            related = getattr(mapping, accessor)
        except Exception:
            continue
        if hasattr(related, "exists"):
            if related.exists():
                return rel.related_model.__name__
        elif related is not None:
            return rel.related_model.__name__
    return None


def assert_can_archive(mapping: StreamDomainMapping):
    blocker = blocking_foreign_key_usage(mapping)
    if blocker:
        raise ValidationError(f"Cannot archive: referenced by {blocker}.")


@transaction.atomic
def create_mapping(*, user, validated_data: dict) -> StreamDomainMapping:
    instance = StreamDomainMapping(**validated_data)
    instance.save(user=user)
    return instance


@transaction.atomic
def update_mapping(*, mapping: StreamDomainMapping, user, validated_data: dict) -> StreamDomainMapping:
    for k, v in validated_data.items():
        setattr(mapping, k, v)
    mapping.save(user=user)
    return mapping


@transaction.atomic
def archive_mapping(*, mapping: StreamDomainMapping, user) -> StreamDomainMapping:
    assert_can_archive(mapping)
    mapping.soft_delete(user=user)
    return mapping


@transaction.atomic
def restore_mapping(*, mapping: StreamDomainMapping, user) -> StreamDomainMapping:
    BaseService.restore(mapping, user=user)
    return mapping


@transaction.atomic
def bulk_archive(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    qs = StreamDomainMapping.objects.filter(id__in=ids, deleted=False)
    count = 0
    for mapping in qs:
        assert_can_archive(mapping)
        mapping.soft_delete(user=user)
        count += 1
    return count


@transaction.atomic
def bulk_restore(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    return StreamDomainMapping.objects.filter(id__in=ids, deleted=True).update(
        deleted=False,
        deleted_at=None,
        deleted_by=None,
        updated_at=timezone.now(),
        updated_by=user,
    )


@transaction.atomic
def set_active_status(*, mapping: StreamDomainMapping, user, is_active: bool) -> StreamDomainMapping:
    mapping.is_active = is_active
    mapping.save(user=user)
    return mapping


@transaction.atomic
def bulk_set_active(*, ids: list, user, is_active: bool) -> int:
    if not ids:
        return 0
    return StreamDomainMapping.objects.filter(id__in=ids).update(
        is_active=is_active,
        updated_at=timezone.now(),
        updated_by=user,
    )


def by_stream_queryset(*, stream_id):
    return (
        StreamDomainMapping.objects.filter(stream_id=stream_id, deleted=False, is_active=True)
        .select_related("domain", "stream")
        .order_by("-weight_score", "domain__domain_name")
    )


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {}
    for k, v in row.items():
        kk = (str(k) if k is not None else "").strip()
        if not kk:
            continue
        key = HEADER_ALIASES.get(kk.lower(), kk.lower())
        out[key] = v

    if "weight_score" in out and out["weight_score"] not in ("", None):
        try:
            out["weight_score"] = int(float(str(out["weight_score"]).strip()))
            BaseService.validate_weight(out["weight_score"])
        except (TypeError, ValueError):
            raise ValueError("Invalid weight_score")
    for b in ("is_primary", "is_active"):
        if b in out:
            out[b] = _to_bool(out.get(b), default=(False if b == "is_primary" else True))
    return out


def parse_import_file(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    if not uploaded:
        return [], ["No file uploaded."]
    name = (getattr(uploaded, "name", "") or "").lower()
    raw = uploaded.read()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return [], ["Excel support requires openpyxl. Install openpyxl or upload CSV."]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], ["Empty spreadsheet."]
            headers = []
            for h in header_row:
                raw_h = str(h).strip() if h is not None else ""
                if raw_h:
                    headers.append(HEADER_ALIASES.get(raw_h.lower(), raw_h.lower()))
                else:
                    headers.append("")
            missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in headers if h]))
            if missing:
                return [], [f"Missing required headers: {', '.join(missing)}"]
            out_rows = []
            for tup in rows_iter:
                if all(x is None or str(x).strip() == "" for x in tup):
                    continue
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = tup[i] if i < len(tup) else None
                out_rows.append(row)
            wb.close()
            return out_rows, errors

        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], ["CSV has no header row."]
        normalized_headers = [
            HEADER_ALIASES.get((str(h).strip().lower() if h else ""), (str(h).strip().lower() if h else ""))
            for h in reader.fieldnames
        ]
        missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in normalized_headers if h]))
        if missing:
            return [], [f"Missing required headers: {', '.join(missing)}"]
        out_rows = []
        for r in reader:
            if not any((v or "").strip() for v in r.values()):
                continue
            out_rows.append(dict(r))
        return out_rows, errors
    except Exception as e:
        logger.exception("parse_import_file failed")
        return [], [str(e)]


@transaction.atomic
def bulk_import_rows(*, user, rows: list[dict], serializer_class, context: dict) -> StreamDomainMappingImportBatch:
    batch = StreamDomainMappingImportBatch.objects.create(created_by=user, total_rows=len(rows))
    imported = 0
    errors: list[StreamDomainMappingImportError] = []
    for idx, raw_row in enumerate(rows, start=1):
        row = dict(raw_row)
        try:
            row = normalize_import_row(row)
        except ValueError as e:
            errors.append(
                StreamDomainMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=dict(raw_row) if isinstance(raw_row, dict) else {},
                )
            )
            continue

        stream = _resolve_stream(row.get("stream_code"))
        if stream is None:
            errors.append(
                StreamDomainMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message="Invalid stream_code (use active stream_code).",
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue

        domain = _resolve_domain(row.get("domain_code"))
        if domain is None:
            errors.append(
                StreamDomainMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message="Invalid domain_code (use active domain_code).",
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue

        payload = {
            "stream": str(stream.pk),
            "domain": str(domain.pk),
            "weight_score": row.get("weight_score"),
            "is_primary": row.get("is_primary", False),
            "is_active": row.get("is_active", True),
        }
        existing = StreamDomainMapping.objects.filter(stream=stream, domain=domain).first()
        ser = serializer_class(instance=existing, data=payload, partial=bool(existing), context=context)
        if not ser.is_valid():
            msg = json.dumps(ser.errors)[:500]
            errors.append(
                StreamDomainMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        try:
            with transaction.atomic():
                obj = ser.save()
                if getattr(obj, "deleted", False):
                    obj.deleted = False
                    obj.deleted_at = None
                    obj.deleted_by = None
                    obj.updated_by = user
                    obj.updated_at = timezone.now()
                    obj.save(update_fields=["deleted", "deleted_at", "deleted_by", "updated_by", "updated_at"])
                imported += 1
        except ValidationError as e:
            detail = e.detail
            msg = json.dumps(detail)[:500] if isinstance(detail, (dict, list)) else str(detail)[:500]
            errors.append(
                StreamDomainMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
        except Exception as e:
            logger.exception("bulk row %s", idx)
            errors.append(
                StreamDomainMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
    if errors:
        StreamDomainMappingImportError.objects.bulk_create(errors)
    batch.imported_count = imported
    batch.failed_count = len(errors)
    batch.completed_at = timezone.now()
    batch.save(update_fields=["imported_count", "failed_count", "completed_at"])
    return batch


def bulk_import_mappings(*, user, rows: list[dict], serializer_class, context: dict) -> dict[str, Any]:
    batch = bulk_import_rows(user=user, rows=rows, serializer_class=serializer_class, context=context)
    err_qs = StreamDomainMappingImportError.objects.filter(batch=batch).order_by("row_number")
    error_details = [{"row": e.row_number, "message": e.message, "row_data": e.row_data} for e in err_qs]
    return {
        "success_count": batch.imported_count,
        "error_count": batch.failed_count,
        "error_details": error_details,
        "batch_id": str(batch.id),
    }


def sample_csv_bytes() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SAMPLE_CSV_HEADERS)
    w.writerow(["science", "engineering", "95", "1", "1"])
    w.writerow(["commerce", "business", "90", "1", "1"])
    return buf.getvalue().encode("utf-8")
