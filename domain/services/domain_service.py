import csv
import io
import json
import logging
from typing import Any
from uuid import UUID

from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from base import services as base_services
from domain.models import Domain, DomainImportBatch, DomainImportError

logger = logging.getLogger(__name__)

SAMPLE_CSV_HEADERS = (
    "domain_code",
    "domain_name",
    "parent",
    "acceptance_level",
    "score",
    "description",
    "is_active",
    "interest_weight",
    "aptitude_weight",
    "personality_weight",
    "work_style_weight",
)


def domain_base_queryset():
    return Domain.objects.select_related("parent", "created_by", "updated_by")


def list_domains(*, include_archived: bool = False):
    qs = domain_base_queryset()
    if not include_archived:
        qs = qs.filter(deleted=False)
    return qs.order_by("-created_at")


def get_domain(*, pk: UUID, include_archived: bool = False):
    qs = domain_base_queryset().filter(pk=pk)
    if not include_archived:
        qs = qs.filter(deleted=False)
    return qs.first()


def filter_domains(
    queryset,
    *,
    is_active=None,
    parent_id=None,
    parent__isnull=None,
):
    if is_active is not None:
        queryset = queryset.filter(is_active=is_active)
    if parent_id is not None:
        queryset = queryset.filter(parent_id=parent_id)
    if parent__isnull is True:
        queryset = queryset.filter(parent__isnull=True)
    elif parent__isnull is False:
        queryset = queryset.filter(parent__isnull=False)
    return queryset


def assert_no_circular_parent(*, domain: Domain | None, parent: Domain | None):
    if parent is None:
        return
    if domain and parent.pk == domain.pk:
        raise ValidationError({"parent_id": "Domain cannot be its own parent."})
    current = parent
    seen = set()
    while current is not None:
        if domain and current.pk == domain.pk:
            raise ValidationError(
                {"parent_id": "Circular parent chain is not allowed."}
            )
        if current.pk in seen:
            raise ValidationError(
                {"parent_id": "Circular parent chain is not allowed."}
            )
        seen.add(current.pk)
        current = current.parent


def blocking_foreign_key_usage(domain: Domain) -> str | None:
    for rel in domain._meta.related_objects:
        if rel.related_model is Domain:
            continue
        accessor = rel.get_accessor_name()
        try:
            related = getattr(domain, accessor)
        except Exception:
            continue
        if hasattr(related, "exists"):
            if related.exists():
                return rel.related_model.__name__
        elif related is not None:
            return rel.related_model.__name__
    return None


def assert_can_archive(domain: Domain):
    if Domain.objects.filter(parent=domain, deleted=False).exists():
        raise ValidationError("Cannot archive: active child domains exist.")
    blocker = blocking_foreign_key_usage(domain)
    if blocker:
        raise ValidationError(f"Cannot archive: referenced by {blocker}.")


def validate_domain_data(
    *,
    data: dict[str, Any],
    instance: Domain | None = None,
    parent: Domain | None = None,
    update_parent: bool = False,
) -> dict[str, Any]:
    code = (data.get("domain_code") or "").strip()
    if not code:
        raise ValidationError({"domain_code": "This field may not be blank."})
    name = (data.get("domain_name") or "").strip()
    if not name:
        raise ValidationError({"domain_name": "This field may not be blank."})
    pal = data.get("parent_acceptance_level")
    if pal is None:
        raise ValidationError({"parent_acceptance_level": "This field is required."})
    pal = int(pal)
    if pal < 1 or pal > 5:
        raise ValidationError({"parent_acceptance_level": "Must be between 1 and 5."})
    score = data.get("future_relevance_score")
    if score is None:
        raise ValidationError({"future_relevance_score": "This field is required."})
    score = int(score)
    if score < 1 or score > 100:
        raise ValidationError({"future_relevance_score": "Must be between 1 and 100."})
    exclude_pk = instance.pk if instance and instance.pk else None
    if case_insensitive_code_exists(code=code, exclude_pk=exclude_pk):
        raise ValidationError(
            {"domain_code": "Domain code must be unique (case-insensitive)."}
        )
    if update_parent or instance is None:
        assert_no_circular_parent(domain=instance, parent=parent)
    return {
        "domain_code": code,
        "domain_name": name,
        "parent_acceptance_level": pal,
        "future_relevance_score": score,
        "description": (data.get("description") or "").strip(),
        "is_active": bool(data.get("is_active", True)),
    }


@transaction.atomic
def create_domain(*, user, validated_data: dict, parent=None) -> Domain:
    assert_no_circular_parent(domain=None, parent=parent)
    instance = Domain(**validated_data, parent=parent)
    instance.save(user=user)
    return instance


@transaction.atomic
def update_domain(
    *,
    domain: Domain,
    user,
    validated_data: dict,
    parent=None,
    update_parent: bool = False,
) -> Domain:
    if update_parent:
        if parent is not None:
            assert_no_circular_parent(domain=domain, parent=parent)
        domain.parent = parent
    for k, v in validated_data.items():
        setattr(domain, k, v)
    domain.save(user=user)
    return domain


@transaction.atomic
def soft_archive_domain(*, domain: Domain, user) -> Domain:
    assert_can_archive(domain)
    return base_services.soft_delete(domain, user=user)


archive_domain = soft_archive_domain


@transaction.atomic
def restore_domain(*, domain: Domain, user) -> Domain:
    return base_services.restore(domain, user=user)


@transaction.atomic
def bulk_archive(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    qs = Domain.objects.filter(id__in=ids, deleted=False)
    count = 0
    for d in qs.select_related("parent"):
        assert_can_archive(d)
        d.soft_delete(user=user)
        count += 1
    return count


@transaction.atomic
def bulk_restore(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    updated = Domain.objects.filter(id__in=ids, deleted=True).update(
        deleted=False,
        deleted_at=None,
        deleted_by=None,
        updated_at=timezone.now(),
        updated_by=user,
    )
    return updated


@transaction.atomic
def set_active_status(*, domain: Domain, user, is_active: bool) -> Domain:
    domain.is_active = is_active
    domain.save(user=user)
    return domain


@transaction.atomic
def bulk_set_active(*, ids: list, user, is_active: bool) -> int:
    if not ids:
        return 0
    return Domain.objects.filter(id__in=ids).update(
        is_active=is_active,
        updated_at=timezone.now(),
        updated_by=user,
    )


def dropdown_domains():
    return (
        Domain.objects.filter(is_active=True, deleted=False)
        .only("id", "domain_code", "domain_name", "parent_id")
        .order_by("domain_name")
    )


def tree_domains():
    rows = list(
        Domain.objects.filter(deleted=False)
        .only(
            "id",
            "domain_code",
            "domain_name",
            "parent_id",
            "is_active",
            "future_relevance_score",
        )
        .order_by("domain_name")
    )
    by_parent: dict[Any, list] = {}
    for r in rows:
        pid = r.parent_id
        by_parent.setdefault(pid, []).append(r)
    return _build_tree(by_parent, None)


def _build_tree(by_parent, parent_id):
    nodes = []
    for obj in by_parent.get(parent_id, []):
        nodes.append(
            {
                "id": str(obj.id),
                "domain_code": obj.domain_code,
                "domain_name": obj.domain_name,
                "is_active": obj.is_active,
                "future_relevance_score": obj.future_relevance_score,
                "children": _build_tree(by_parent, obj.id),
            }
        )
    return nodes


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {}
    for k, v in row.items():
        kk = (str(k) if k is not None else "").strip()
        if not kk:
            continue
        out[kk] = v
    if "acceptance_level" in out and "parent_acceptance_level" not in out:
        out["parent_acceptance_level"] = out.pop("acceptance_level")
    if "score" in out and "future_relevance_score" not in out:
        out["future_relevance_score"] = out.pop("score")
    for k in ("parent_acceptance_level", "future_relevance_score"):
        if k in out and out[k] not in ("", None):
            if isinstance(out[k], int):
                continue
            try:
                out[k] = int(float(str(out[k]).strip()))
            except (TypeError, ValueError):
                raise ValueError(f"Invalid {k}")
    if "is_active" in out and out["is_active"] not in ("", None):
        v = str(out["is_active"]).lower()
        out["is_active"] = v in ("1", "true", "yes", "y")

    # Optional affinity weights (floats 0..1). Keep None/"" as None.
    for k in ("interest_weight", "aptitude_weight", "personality_weight", "work_style_weight"):
        if k in out and out[k] not in ("", None):
            try:
                out[k] = float(str(out[k]).strip())
            except (TypeError, ValueError):
                raise ValueError(f"Invalid {k}")
    return out


def resolve_parent_code_in_row(row: dict[str, Any]) -> str | None:
    if row.get("parent_id") not in (None, ""):
        return None
    raw = row.get("parent")
    if raw is None:
        raw = row.get("parent_code")
    if raw is None or str(raw).strip() == "":
        row.pop("parent", None)
        row.pop("parent_code", None)
        return None
    pc = str(raw).strip()
    row.pop("parent", None)
    row.pop("parent_code", None)
    p = Domain.objects.filter(domain_code__iexact=pc, deleted=False).first()
    if not p:
        return f"Unknown parent domain_code: {pc}"
    row["parent_id"] = str(p.id)
    return None


def parse_import_file(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    if not uploaded:
        return [], ["No file uploaded."]
    name = (getattr(uploaded, "name", "") or "").lower()
    raw = uploaded.read()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return [], [
                    "Excel support requires openpyxl. Install openpyxl or upload CSV."
                ]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], ["Empty spreadsheet."]
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            out_rows = []
            for tup in rows_iter:
                if all(x is None or str(x).strip() == "" for x in tup):
                    continue
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = tup[i] if i < len(tup) else None
                out_rows.append(row)
            wb.close()
            return out_rows, errors
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], ["CSV has no header row."]
        out_rows = []
        for r in reader:
            if not any((v or "").strip() for v in r.values()):
                continue
            out_rows.append(dict(r))
        return out_rows, errors
    except Exception as e:
        logger.exception("parse_import_file failed")
        return [], [str(e)]


def bulk_import_rows(
    *, user, rows: list[dict], serializer_class, context: dict
) -> DomainImportBatch:
    batch = DomainImportBatch.objects.create(
        created_by=user,
        total_rows=len(rows),
    )
    imported = 0
    errors: list[DomainImportError] = []
    seen_codes: set[str] = set()
    for idx, raw_row in enumerate(rows, start=1):
        row = dict(raw_row)
        try:
            row = normalize_import_row(row)
        except ValueError as e:
            errors.append(
                DomainImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=dict(raw_row) if isinstance(raw_row, dict) else {},
                )
            )
            continue
        row_code = (row.get("domain_code") or "").strip().lower()
        if row_code:
            if row_code in seen_codes:
                errors.append(
                    DomainImportError(
                        batch=batch,
                        row_number=idx,
                        message=f"Duplicate domain_code in upload: {row_code}"[:500],
                        row_data=row if isinstance(row, dict) else {},
                    )
                )
                continue
            seen_codes.add(row_code)
        perr = resolve_parent_code_in_row(row)
        if perr:
            errors.append(
                DomainImportError(
                    batch=batch,
                    row_number=idx,
                    message=perr[:500],
                    row_data=dict(raw_row) if isinstance(raw_row, dict) else {},
                )
            )
            continue
        existing = None
        domain_code = (row.get("domain_code") or "").strip()
        if domain_code:
            existing = Domain.objects.filter(domain_code__iexact=domain_code).first()
        ser = serializer_class(
            instance=existing, data=row, partial=bool(existing), context=context
        )
        if not ser.is_valid():
            msg = json.dumps(ser.errors)[:500]
            errors.append(
                DomainImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        try:
            with transaction.atomic():
                obj = ser.save()
                # Restore archived records when re-imported.
                if getattr(obj, "deleted", False):
                    obj.deleted = False
                    obj.deleted_at = None
                    obj.deleted_by = None
                    obj.updated_by = user
                    obj.updated_at = timezone.now()
                    obj.save(
                        update_fields=[
                            "deleted",
                            "deleted_at",
                            "deleted_by",
                            "updated_by",
                            "updated_at",
                        ]
                    )
                imported += 1
        except ValidationError as e:
            detail = e.detail
            if isinstance(detail, (dict, list)):
                msg = json.dumps(detail)[:500]
            else:
                msg = str(detail)[:500]
            errors.append(
                DomainImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
        except Exception as e:
            logger.exception("bulk row %s", idx)
            errors.append(
                DomainImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
    if errors:
        DomainImportError.objects.bulk_create(errors)
    batch.imported_count = imported
    batch.failed_count = len(errors)
    batch.completed_at = timezone.now()
    batch.save(
        update_fields=["imported_count", "failed_count", "completed_at"],
    )
    return batch


def bulk_import_domains(
    *, user, rows: list[dict], serializer_class, context: dict
) -> dict[str, Any]:
    batch = bulk_import_rows(
        user=user,
        rows=rows,
        serializer_class=serializer_class,
        context=context,
    )
    err_qs = DomainImportError.objects.filter(batch=batch).order_by("row_number")
    error_details = [
        {"row": e.row_number, "message": e.message, "row_data": e.row_data}
        for e in err_qs.iterator(chunk_size=200)
    ]
    return {
        "success_count": batch.imported_count,
        "error_count": batch.failed_count,
        "error_details": error_details,
        "batch_id": str(batch.id),
    }


def import_batches_queryset():
    return DomainImportBatch.objects.select_related("created_by").order_by(
        "-created_at"
    )


def import_errors_queryset(*, batch_id: UUID | None = None):
    qs = DomainImportError.objects.select_related("batch").order_by(
        "-batch__created_at", "row_number"
    )
    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    return qs


def error_report_csv_bytes(*, batch_id: UUID | None = None) -> tuple[str, bytes]:
    qs = import_errors_queryset(batch_id=batch_id)
    if batch_id is None:
        first = qs.first()
        if not first:
            return "domain_import_errors_empty.csv", b"row_number,message\n"
        qs = import_errors_queryset(batch_id=first.batch_id)
        filename = f"domain_import_errors_{first.batch_id}.csv"
    else:
        filename = f"domain_import_errors_{batch_id}.csv"
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["row_number", "message", "row_data"])
    for er in qs.iterator(chunk_size=200):
        w.writerow([er.row_number, er.message, er.row_data])
    return filename, buf.getvalue().encode("utf-8")


def case_insensitive_code_exists(*, code: str, exclude_pk: UUID | None = None) -> bool:
    q = Domain.objects.filter(domain_code__iexact=code)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    return q.exists()


def sample_csv_bytes() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SAMPLE_CSV_HEADERS)
    w.writerow(["ROOT", "Root domain", "", "3", "80", "Top-level sample", "1", "", "", "", ""])
    w.writerow(["CHILD_A", "Child A", "ROOT", "2", "60", "Under ROOT", "1", "0.25", "0.25", "0.25", "0.25"])
    return buf.getvalue().encode("utf-8")
