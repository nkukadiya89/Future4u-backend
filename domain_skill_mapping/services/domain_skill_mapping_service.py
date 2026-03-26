import csv
import io
import json
import logging
from typing import Any
from uuid import UUID

from django.db import transaction
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from base.services import BaseService
from domain_skill_mapping.models import (
    DomainSkillMapping,
    DomainSkillMappingImportBatch,
    DomainSkillMappingImportError,
)

logger = logging.getLogger(__name__)

SAMPLE_CSV_HEADERS = ("domain_code", "skill_code", "weight_score", "is_core", "is_active")
REQUIRED_IMPORT_HEADERS = {"domain_code", "skill_code", "weight_score"}
HEADER_ALIASES = {
    "domain": "domain_code",
    "skill": "skill_code",
    "score": "weight_score",
    "core": "is_core",
    "active": "is_active",
}


def mapping_base_queryset():
    return DomainSkillMapping.objects.select_related("domain", "skill", "created_by", "updated_by")


def _to_bool(value, *, default=False):
    if value in (None, ""):
        return default
    return str(value).strip().lower() in ("1", "true", "yes", "y")


def _resolve_domain(domain_code):
    if domain_code in (None, ""):
        return None
    from domain.models import Domain

    return Domain.objects.filter(domain_code__iexact=str(domain_code).strip(), deleted=False).first()


def _resolve_skill(skill_code):
    if skill_code in (None, ""):
        return None
    from skill.models import Skill

    return Skill.objects.filter(skill_code__iexact=str(skill_code).strip(), deleted=False).first()


def pair_exists(*, domain_id, skill_id, exclude_pk: UUID | None = None) -> bool:
    q = DomainSkillMapping.objects.filter(domain_id=domain_id, skill_id=skill_id)
    if exclude_pk:
        q = q.exclude(pk=exclude_pk)
    return q.exists()


def blocking_foreign_key_usage(mapping: DomainSkillMapping) -> str | None:
    for rel in mapping._meta.related_objects:
        accessor = rel.get_accessor_name()
        try:
            related = getattr(mapping, accessor)
        except Exception:
            continue
        if hasattr(related, "exists"):
            if related.exists():
                return rel.related_model.__name__
        elif related is not None:
            return rel.related_model.__name__
    return None


def assert_can_archive(mapping: DomainSkillMapping):
    blocker = blocking_foreign_key_usage(mapping)
    if blocker:
        raise ValidationError(f"Cannot archive: referenced by {blocker}.")


@transaction.atomic
def create_mapping(*, user, validated_data: dict) -> DomainSkillMapping:
    instance = DomainSkillMapping(**validated_data)
    instance.save(user=user)
    return instance


@transaction.atomic
def update_mapping(*, mapping: DomainSkillMapping, user, validated_data: dict) -> DomainSkillMapping:
    for k, v in validated_data.items():
        setattr(mapping, k, v)
    mapping.save(user=user)
    return mapping


@transaction.atomic
def archive_mapping(*, mapping: DomainSkillMapping, user) -> DomainSkillMapping:
    assert_can_archive(mapping)
    mapping.soft_delete(user=user)
    return mapping


@transaction.atomic
def restore_mapping(*, mapping: DomainSkillMapping, user) -> DomainSkillMapping:
    BaseService.restore(mapping, user=user)
    return mapping


@transaction.atomic
def bulk_archive(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    qs = DomainSkillMapping.objects.filter(id__in=ids, deleted=False)
    count = 0
    for mapping in qs:
        assert_can_archive(mapping)
        mapping.soft_delete(user=user)
        count += 1
    return count


@transaction.atomic
def bulk_restore(*, ids: list, user) -> int:
    if not ids:
        raise ValidationError({"ids": "This field is required."})
    return DomainSkillMapping.objects.filter(id__in=ids, deleted=True).update(
        deleted=False,
        deleted_at=None,
        deleted_by=None,
        updated_at=timezone.now(),
        updated_by=user,
    )


@transaction.atomic
def set_active_status(*, mapping: DomainSkillMapping, user, is_active: bool) -> DomainSkillMapping:
    mapping.is_active = is_active
    mapping.save(user=user)
    return mapping


@transaction.atomic
def bulk_set_active(*, ids: list, user, is_active: bool) -> int:
    if not ids:
        return 0
    return DomainSkillMapping.objects.filter(id__in=ids).update(
        is_active=is_active,
        updated_at=timezone.now(),
        updated_by=user,
    )


def by_domain_queryset(*, domain_id):
    return (
        DomainSkillMapping.objects.filter(domain_id=domain_id, deleted=False, is_active=True)
        .select_related("domain", "skill")
        .order_by("-weight_score", "skill__skill_name")
    )


def normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {}
    for k, v in row.items():
        kk = (str(k) if k is not None else "").strip()
        if not kk:
            continue
        key = HEADER_ALIASES.get(kk.lower(), kk.lower())
        out[key] = v

    if "weight_score" in out and out["weight_score"] not in ("", None):
        try:
            out["weight_score"] = int(float(str(out["weight_score"]).strip()))
            BaseService.validate_weight(out["weight_score"])
        except (TypeError, ValueError):
            raise ValueError("Invalid weight_score")
    for b in ("is_core", "is_active"):
        if b in out:
            out[b] = _to_bool(out.get(b), default=(False if b == "is_core" else True))
    return out


def parse_import_file(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    if not uploaded:
        return [], ["No file uploaded."]
    name = (getattr(uploaded, "name", "") or "").lower()
    raw = uploaded.read()
    try:
        if name.endswith(".xlsx") or name.endswith(".xls"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return [], ["Excel support requires openpyxl. Install openpyxl or upload CSV."]
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return [], ["Empty spreadsheet."]
            headers = []
            for h in header_row:
                raw_h = str(h).strip() if h is not None else ""
                if raw_h:
                    headers.append(HEADER_ALIASES.get(raw_h.lower(), raw_h.lower()))
                else:
                    headers.append("")
            missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in headers if h]))
            if missing:
                return [], [f"Missing required headers: {', '.join(missing)}"]
            out_rows = []
            for tup in rows_iter:
                if all(x is None or str(x).strip() == "" for x in tup):
                    continue
                row = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    row[h] = tup[i] if i < len(tup) else None
                out_rows.append(row)
            wb.close()
            return out_rows, errors

        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            return [], ["CSV has no header row."]
        normalized_headers = [
            HEADER_ALIASES.get((str(h).strip().lower() if h else ""), (str(h).strip().lower() if h else ""))
            for h in reader.fieldnames
        ]
        missing = sorted(REQUIRED_IMPORT_HEADERS - set([h for h in normalized_headers if h]))
        if missing:
            return [], [f"Missing required headers: {', '.join(missing)}"]
        out_rows = []
        for r in reader:
            if not any((v or "").strip() for v in r.values()):
                continue
            out_rows.append(dict(r))
        return out_rows, errors
    except Exception as e:
        logger.exception("parse_import_file failed")
        return [], [str(e)]


@transaction.atomic
def bulk_import_rows(*, user, rows: list[dict], serializer_class, context: dict) -> DomainSkillMappingImportBatch:
    batch = DomainSkillMappingImportBatch.objects.create(created_by=user, total_rows=len(rows))
    imported = 0
    errors: list[DomainSkillMappingImportError] = []
    for idx, raw_row in enumerate(rows, start=1):
        row = dict(raw_row)
        try:
            row = normalize_import_row(row)
        except ValueError as e:
            errors.append(
                DomainSkillMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=dict(raw_row) if isinstance(raw_row, dict) else {},
                )
            )
            continue

        domain = _resolve_domain(row.get("domain_code"))
        if domain is None:
            errors.append(
                DomainSkillMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message="Invalid domain_code (use active domain_code).",
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue

        skill = _resolve_skill(row.get("skill_code"))
        if skill is None:
            errors.append(
                DomainSkillMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message="Invalid skill_code (use active skill_code).",
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue

        payload = {
            "domain": str(domain.pk),
            "skill": str(skill.pk),
            "weight_score": row.get("weight_score"),
            "is_core": row.get("is_core", False),
            "is_active": row.get("is_active", True),
        }
        existing = DomainSkillMapping.objects.filter(domain=domain, skill=skill).first()
        ser = serializer_class(instance=existing, data=payload, partial=bool(existing), context=context)
        if not ser.is_valid():
            msg = json.dumps(ser.errors)[:500]
            errors.append(
                DomainSkillMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
            continue
        try:
            with transaction.atomic():
                obj = ser.save()
                if getattr(obj, "deleted", False):
                    obj.deleted = False
                    obj.deleted_at = None
                    obj.deleted_by = None
                    obj.updated_by = user
                    obj.updated_at = timezone.now()
                    obj.save(update_fields=["deleted", "deleted_at", "deleted_by", "updated_by", "updated_at"])
                imported += 1
        except ValidationError as e:
            detail = e.detail
            msg = json.dumps(detail)[:500] if isinstance(detail, (dict, list)) else str(detail)[:500]
            errors.append(
                DomainSkillMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=msg,
                    row_data=row if isinstance(row, dict) else {},
                )
            )
        except Exception as e:
            logger.exception("bulk row %s", idx)
            errors.append(
                DomainSkillMappingImportError(
                    batch=batch,
                    row_number=idx,
                    message=str(e)[:500],
                    row_data=row if isinstance(row, dict) else {},
                )
            )
    if errors:
        DomainSkillMappingImportError.objects.bulk_create(errors)
    batch.imported_count = imported
    batch.failed_count = len(errors)
    batch.completed_at = timezone.now()
    batch.save(update_fields=["imported_count", "failed_count", "completed_at"])
    return batch


def bulk_import_mappings(*, user, rows: list[dict], serializer_class, context: dict) -> dict[str, Any]:
    batch = bulk_import_rows(user=user, rows=rows, serializer_class=serializer_class, context=context)
    err_qs = DomainSkillMappingImportError.objects.filter(batch=batch).order_by("row_number")
    error_details = [{"row": e.row_number, "message": e.message, "row_data": e.row_data} for e in err_qs]
    return {
        "success_count": batch.imported_count,
        "error_count": batch.failed_count,
        "error_details": error_details,
        "batch_id": str(batch.id),
    }


def sample_csv_bytes() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SAMPLE_CSV_HEADERS)
    w.writerow(["engineering", "python", "95", "1", "1"])
    w.writerow(["business", "excel", "90", "0", "1"])
    return buf.getvalue().encode("utf-8")

